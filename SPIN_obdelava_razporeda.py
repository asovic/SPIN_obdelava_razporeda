import openpyxl
try:
    wb = openpyxl.load_workbook('SPIN_izvajalci_razpored_razpolozljivost.xlsx')
    print("Datoteka najdena.\nBerem...")
except FileNotFoundError:
    print("Datoteka 'SPIN_izvajalci_razpored_razpolozljivost.xlsx' ni najdena.")
    print("\nPreveri ime datoteke oz. če se datoteka in program nahajata v isti mapi.")
sheet = wb.active

class Gather_Data:
    celice_datumov = sheet['B2':'B22']
    celice_izvajalcev = sheet['C1':'W1']
    def __init__(self):
        self.datumi = []
        self.lista_izvajalcev = []

    def datum(self):
            for date in self.celice_datumov:
                for dates in date:
                    self.datumi.append(dates.value)
            return self.datumi

    def izvajalci(self):
            for iz in self.celice_izvajalcev:
                for izv in iz:
                    self.lista_izvajalcev.append(izv.value)
            return self.lista_izvajalcev

    def urnik(self,i):
        ure = []
        start_cell = 'C{}'.format(i+2)
        end_cell = 'W{}'.format(i+2)
        cell = sheet[start_cell:end_cell]
        for c1 in cell:
            for c_val in c1:
                ure.append(c_val.value)
        return ure

data = Gather_Data()

def create_tripple(datumi, ure): #Generator trojcka, ce je razpolozljivost v celici podana
    for i, v in enumerate(ure):
        if v is not None:
            yield datumi, ure[i], data.izvajalci()[i]

def prewrite_list(): #Seznam iz tupla datuma, ure in izvajalca 
    m_list = []
    for i, v in enumerate(data.datum()):
        one = list(create_tripple(v, data.urnik(i)))
        m_list.append(one)
    return m_list

#def sort_list_by_time(input_list):
#    for i in range(0, len(input_list)):
#        if i >= 2:
#            input_list[i].sort(key = lambda x: x[1][0], reverse = True)
#        return input_list
#
#sorted_list = sort_list_by_time(prewrite_list)

def write_list(): #Pretvori iz nested lista v list s tupli
    w_list = []
    for group_date in prewrite_list():
        for single_date in group_date:
            w_list.append(single_date)
    return w_list

#Open xlsx file for writing
wb_write = openpyxl.Workbook()#openpyxl.load_workbook('Book1.xlsx')
sheet_w = wb_write.active

def write_to_xls(): #Zapis podatkov v celice zvezka
    for i, v in enumerate(write_list()):
        date, time, izv = v
        A_cell = 'A{}'.format(i+1)
        B_cell = 'B{}'.format(i+1)
        C_cell = 'C{}'.format(i+1)
        date_cell = sheet_w[A_cell]
        time_cell = sheet_w[B_cell]
        izv_cell = sheet_w[C_cell]
        date_cell.value = date
        time_cell.value = time
        izv_cell.value = izv
filename = input("Vnesi ime nove datoteke: ") +".xlsx"
write_to_xls()
wb_write.save(filename)
print("Nova datoteka ustvarjena.")
print("© Andrej Sovič\nV primeru težav: asovic@me.com")
input()